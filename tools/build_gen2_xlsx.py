#!/usr/bin/env python3
"""Genera docs/FE4_Gen2_Inheritance.xlsx a partir de gen2_wod.json.
Una hoja por personaje: máximos, base (mínimos), rango de growths y el
desglose completo por padre (Nv.1 / Nv.30 / growth%)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = os.path.join(os.path.dirname(__file__), "gen2_wod.json")
OUT = os.path.join(os.path.dirname(__file__), "..", "..", "..", "..",
                   "home", "user", "FireEmblem4_5_Remake", "docs",
                   "FE4_Gen2_Inheritance.xlsx")
# Resolver ruta de salida de forma robusta: repo root conocido.
REPO = "/home/user/FireEmblem4_5_Remake"
OUT = os.path.join(REPO, "docs", "FE4_Gen2_Inheritance.xlsx")

STATS = ["HP", "STR", "MAG", "SKL", "SPD", "LCK", "DEF", "RES"]

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="374785")
SUB_FILL = PatternFill("solid", fgColor="A8B4E0")
MAXF = PatternFill("solid", fgColor="F5D76E")
BASEF = PatternFill("solid", fgColor="C8E6C9")
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center")


def _row(ws, r, cells, fill=None, bold=False):
    for c, val in enumerate(cells, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = BOX
        cell.alignment = CEN
        if fill:
            cell.fill = fill
        if bold:
            cell.font = Font(bold=True)
    return r + 1


def _stat_cells(d):
    return [d.get(s, "") for s in STATS]


def build(data):
    wb = Workbook()
    wb.remove(wb.active)
    for nid, ch in data.items():
        if not isinstance(ch, dict) or "fathers" not in ch:
            ws = wb.create_sheet(nid[:31])
            ws["A1"] = "SIN DATOS (%s)" % ch.get("error", "?")
            continue
        ws = wb.create_sheet(nid[:31])
        ws.column_dimensions["A"].width = 16
        for col in "BCDEFGHI":
            ws.column_dimensions[col].width = 7

        r = 1
        # Cabecera de info
        ws.cell(row=r, column=1, value=nid).font = Font(bold=True, size=14)
        r += 1
        for label, key in [("Clase", "klass"), ("Sangre", "holy_blood"),
                           ("Skills pers.", "personal_skills"),
                           ("Items inic.", "starting_items")]:
            v = ch.get(key, "")
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws.cell(row=r, column=2, value=v)
            r += 1
        r += 1

        # Fila de encabezado de stats
        r = _row(ws, r, ["", *STATS], fill=HDR_FILL, bold=True)
        for c in range(1, len(STATS) + 2):
            ws.cell(row=r - 1, column=c).font = HDR

        # Máximos (cap de clase)
        r = _row(ws, r, ["MÁXIMOS", *_stat_cells(ch.get("max_stats", {}))], fill=MAXF)

        # Base = mínimos por stat a Nv.1 entre todos los padres
        fathers = ch.get("fathers", {})
        base = {}
        gmin = {}
        gmax = {}
        for s in STATS:
            l1vals = [f["lv1"].get(s) for f in fathers.values()
                      if isinstance(f, dict) and "lv1" in f and f["lv1"].get(s) is not None]
            gvals = [f["growth"].get(s) for f in fathers.values()
                     if isinstance(f, dict) and "growth" in f and f["growth"].get(s) is not None]
            base[s] = min(l1vals) if l1vals else ""
            gmin[s] = min(gvals) if gvals else ""
            gmax[s] = max(gvals) if gvals else ""
        r = _row(ws, r, ["BASE (mín Nv.1)", *_stat_cells(base)], fill=BASEF)
        r = _row(ws, r, ["GROWTH mín %", *_stat_cells(gmin)])
        r = _row(ws, r, ["GROWTH máx %", *_stat_cells(gmax)])
        r += 1

        # Desglose por padre
        r = _row(ws, r, ["POR PADRE ↓", *STATS], fill=SUB_FILL, bold=True)
        for fname, f in fathers.items():
            if not isinstance(f, dict):
                continue
            sk = ", ".join(f.get("skills", []))
            note = f.get("holy_blood_note", "")
            tag = fname + (" *" if note else "")
            ws.cell(row=r, column=1, value=tag).font = Font(bold=True, italic=True)
            r += 1
            r = _row(ws, r, ["  Nv.1", *_stat_cells(f.get("lv1", {}))])
            r = _row(ws, r, ["  Nv.30", *_stat_cells(f.get("lv30", {}))])
            r = _row(ws, r, ["  Growth%", *_stat_cells(f.get("growth", {}))])
            info = "  skills: " + sk + (("  | " + note) if note else "")
            ws.cell(row=r, column=1, value=info).font = Font(italic=True, size=9)
            r += 2

    wb.save(OUT)
    print("wrote", OUT, "sheets:", len(wb.sheetnames))


if __name__ == "__main__":
    with open(SRC) as fh:
        build(json.load(fh))
