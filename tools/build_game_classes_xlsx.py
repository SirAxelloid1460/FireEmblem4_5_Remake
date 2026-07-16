#!/usr/bin/env python3
"""Genera docs/FE4_FE5_Classes_Real.xlsx desde
data/general/game_classes_wod_source.json (cosecha de fireemblemwod).
Una hoja por juego (FE4, FE5). FE4 con stats completos; FE5 solo nombres/
armas/promoción (la página de clases FE5 está sin rellenar en la fuente)."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

REPO = "/home/user/FireEmblem4_5_Remake"
SRC = os.path.join(REPO, "data", "general", "game_classes_wod_source.json")
OUT = os.path.join(REPO, "docs", "FE4_FE5_Classes_Real.xlsx")

STATS = ["HP", "STR", "MAG", "SKL", "SPD", "LCK", "DEF", "RES", "CON", "MOV"]
CAP_STATS = ["HP", "STR", "MAG", "SKL", "SPD", "LCK", "DEF", "RES"]

HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="374785")
GRP = PatternFill("solid", fgColor="A8B4E0")
THIN = Side(style="thin", color="C0C0C0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CEN = Alignment(horizontal="center")


def _cell(ws, r, c, v, fill=None, bold=False):
    x = ws.cell(row=r, column=c, value=v)
    x.border = BOX
    x.alignment = CEN
    if fill:
        x.fill = fill
    if bold:
        x.font = Font(bold=True)
    return x


def build(classes):
    wb = Workbook()
    wb.remove(wb.active)
    for game in ["FE4", "FE5"]:
        ws = wb.create_sheet(game)
        cs = [c for c in classes if c.get("game") == game]
        # cabecera de grupos (fila 1) y de columnas (fila 2)
        cols = ["name", "category"] + ["b." + s for s in STATS] \
            + ["p." + s for s in STATS] + ["cap." + s for s in CAP_STATS] \
            + ["weapons", "promotes_to", "special"]
        for i, name in enumerate(cols, 1):
            c = _cell(ws, 1, i, name, fill=HFILL, bold=True)
            c.font = HDR
        r = 2
        for cl in cs:
            b = cl.get("bases", {}) or {}
            p = cl.get("promotion_bonuses", {}) or {}
            cap = cl.get("caps", {}) or {}
            w = cl.get("weapons", {}) or {}
            wtxt = ", ".join("%s(%s)" % (k, v) for k, v in w.items())
            row = [cl.get("name", ""), cl.get("category", "")]
            row += [b.get(s, "") for s in STATS]
            row += [p.get(s, "") for s in STATS]
            row += [cap.get(s, "") for s in CAP_STATS]
            row += [wtxt, cl.get("promotes_to", ""), cl.get("special_ability", "")]
            for i, v in enumerate(row, 1):
                _cell(ws, r, i, v)
            r += 1
        # anchos
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 15
        for col in range(3, 3 + len(STATS) * 2 + len(CAP_STATS)):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 5
        for off in range(3):
            ws.column_dimensions[ws.cell(row=1, column=len(cols) - off).column_letter].width = 30
        ws.freeze_panes = "C2"
    wb.save(OUT)
    print("wrote", OUT, "sheets:", wb.sheetnames)


if __name__ == "__main__":
    d = json.load(open(SRC))
    build(d["classes"] if isinstance(d, dict) else d)
