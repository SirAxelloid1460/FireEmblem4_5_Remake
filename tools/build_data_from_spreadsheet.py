#!/usr/bin/env python3
"""
build_data_from_spreadsheet.py
Convierte "Genealogy + Thracia Data.xlsx" en datos nativos de Godot (JSON) bajo
res://data/general/. Es una herramienta OFFLINE (no se usa en runtime); el juego
sólo lee los .json resultantes vía GameDatabase. Reemplaza al antiguo LTImporter.

Uso:  python tools/build_data_from_spreadsheet.py "<ruta al .xlsx>"
"""
import sys, json, os, re

import openpyxl

# Orden REAL de las secciones en la hoja Weapons (verificado por la 1ª arma de
# cada sección). "Magic" agrupa Anima/Light/Dark (refinable por tomo); "Scroll"
# son los pergaminos de FE5 que modifican growths.
WEAPON_TYPES_ORDER = ["Sword", "Lance", "Axe", "Bow", "Ballista", "Magic", "Staff", "Scroll"]
CLASS_WEXP_COLS = ["Sword", "Lance", "Axe", "Bow", "Anima", "Dark", "Holy", "Staff"]
STAT_KEYS = ["hp", "str", "mag", "skl", "spd", "lck", "def", "res"]

def clean(s):
    if s is None:
        return ""
    s = str(s).strip()
    # Sanea comillas curly y el replacement char del xlsx
    s = (s.replace("�", "'").replace("’", "'").replace("‘", "'")
           .replace("“", '"').replace("”", '"').replace("–", "-")
           .replace("—", "-"))
    return s

def num(v, default=0):
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return default

def split_base_cap(cell):
    """'12/60' -> (12, 60). '--' o vacío -> (0,0)."""
    s = clean(cell)
    if "/" in s:
        a, b = s.split("/", 1)
        return num(a), num(b)
    return num(s), num(s)

def parse_classes(ws):
    out = []
    tier = 0
    for r in range(1, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=2).value)
        if not name:
            continue
        up = name.upper()
        if up.startswith("TIER"):
            m = re.search(r"\d+", up)
            tier = int(m.group()) if m else tier
            continue
        if name == "Class Name":
            continue
        bases, caps = [], []
        for col in range(3, 11):  # HP..Res
            b, c = split_base_cap(ws.cell(row=r, column=col).value)
            bases.append(b); caps.append(c)
        mov = num(ws.cell(row=r, column=11).value)
        con = num(ws.cell(row=r, column=12).value)
        wexp = {}
        for i, t in enumerate(CLASS_WEXP_COLS):
            val = clean(ws.cell(row=r, column=13 + i).value)
            if val and val != "--":
                wexp[t] = val
        skills_raw = clean(ws.cell(row=r, column=21).value)
        skills = [s.strip() for s in re.split(r"[,/]", skills_raw) if s.strip() and s.strip() != "--"]
        enemy_only = clean(ws.cell(row=r, column=22).value) not in ("", "--")
        growths = [num(ws.cell(row=r, column=23 + i).value) for i in range(8)]
        # growths vienen como fracciones (0.85) -> a porcentaje entero
        growths = [int(round(g * 100)) if isinstance(g, float) and g <= 1.5 else num(g) for g in growths]
        out.append({
            "id": name, "name": name, "tier": tier,
            "bases": dict(zip(STAT_KEYS, bases)),
            "caps": dict(zip(STAT_KEYS, caps)),
            "growths": dict(zip(STAT_KEYS, growths)),
            "mov": mov, "con": con,
            "wexp": wexp, "skills": skills, "enemy_only": enemy_only,
        })
    return out

def parse_weapons(ws):
    out = []
    type_idx = -1
    prev_blank = True  # arrancamos como si vinieramos de separador
    for r in range(1, ws.max_row + 1):
        c2 = clean(ws.cell(row=r, column=2).value)
        rowvals = [ws.cell(row=r, column=c).value for c in range(2, 11)]
        is_blank = all(v is None or str(v).strip() == "" for v in rowvals)
        if is_blank:
            prev_blank = True
            continue
        if c2 == "Name":  # cabecera de seccion -> nuevo tipo de arma
            if prev_blank or type_idx < 0:
                type_idx += 1
            prev_blank = False
            continue
        prev_blank = False
        if type_idx < 0:
            type_idx = 0
        wtype = WEAPON_TYPES_ORDER[type_idx] if type_idx < len(WEAPON_TYPES_ORDER) else "Unknown"
        rng = clean(ws.cell(row=r, column=6).value)
        if "~" in rng or "-" in rng:
            parts = re.split(r"[~\-]", rng)
            rmin, rmax = num(parts[0], 1), num(parts[-1], 1)
        else:
            rmin = rmax = num(rng, 1)
        rank = clean(ws.cell(row=r, column=7).value)
        worth_raw = clean(ws.cell(row=r, column=8).value)
        out.append({
            "name": c2, "weapon_type": wtype,
            "might": num(ws.cell(row=r, column=3).value),
            "weight": num(ws.cell(row=r, column=4).value),
            "hit": num(ws.cell(row=r, column=5).value),
            "rng_min": rmin, "rng_max": rmax,
            "rank": rank if rank else "E",
            "worth": num(worth_raw, -1),         # -1 = no comprable
            "uses": num(ws.cell(row=r, column=9).value),
            "effects": clean(ws.cell(row=r, column=10).value),
            "personal": rank == "*",             # arma personal/sagrada
        })
    return out

def parse_skills(ws):
    """Skills: col2=Name, col3=Effect, col4=Activation. Salta cabeceras 'Name'."""
    out = []
    seen = set()
    for r in range(1, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=2).value)
        if not name or name == "Name":
            continue
        if name in seen:
            continue
        seen.add(name)
        out.append({
            "name": name,
            "effect": clean(ws.cell(row=r, column=3).value),
            "activation": clean(ws.cell(row=r, column=4).value),
        })
    return out

def parse_items(ws):
    """Items: col3=Name, col4=Worth, col5=Uses, col6=Effect. Las filas con sólo
    col3 (sin worth) se tratan como cabeceras de categoría."""
    out = []
    category = ""
    seen = set()
    for r in range(1, ws.max_row + 1):
        name = clean(ws.cell(row=r, column=3).value)
        if not name or name in ("Name", "Item"):
            continue
        worth_raw = clean(ws.cell(row=r, column=4).value)
        uses_raw = clean(ws.cell(row=r, column=5).value)
        effect = clean(ws.cell(row=r, column=6).value)
        # Cabecera de categoría: nombre presente pero sin worth ni efecto.
        if worth_raw == "" and effect == "":
            category = name
            continue
        if name in seen:
            continue
        seen.add(name)
        out.append({
            "name": name,
            "category": category,
            "worth": num(worth_raw, -1),
            "uses": num(uses_raw, -1) if uses_raw not in ("--", "") else -1,
            "effect": effect,
        })
    return out

def parse_two_col(ws, start_row=3):
    """Skills / cosas con Name|Effect|Activation o Name|...|Effect.
    Devuelve filas crudas a partir de start_row para inspeccion."""
    rows = []
    for r in range(start_row, ws.max_row + 1):
        vals = [clean(ws.cell(row=r, column=c).value) for c in range(1, 7)]
        if any(vals):
            rows.append((r, vals))
    return rows

def main():
    path = sys.argv[1] if len(sys.argv) > 1 else "Genealogy + Thracia Data.xlsx"
    outdir = sys.argv[2] if len(sys.argv) > 2 else "FE4_FE5_Godot/data/general"
    wb = openpyxl.load_workbook(path, data_only=True)
    os.makedirs(outdir, exist_ok=True)

    classes = parse_classes(wb["Classes"])
    weapons = parse_weapons(wb["Weapons"])
    skills = parse_skills(wb["Skills"])
    items = parse_items(wb["Items"])

    def dump(name, data):
        with open(os.path.join(outdir, name), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)

    dump("classes.json", classes)
    dump("weapons.json", weapons)
    dump("skills.json", skills)
    dump("items.json", items)

    print(f"classes: {len(classes)}  -> classes.json")
    print(f"weapons: {len(weapons)} -> weapons.json")
    print(f"skills:  {len(skills)} -> skills.json")
    print(f"items:   {len(items)} -> items.json")
    if skills: print("  sample skill:", json.dumps(skills[1], ensure_ascii=False)[:140])
    if items:  print("  sample item: ", json.dumps(items[0], ensure_ascii=False)[:140])
    # Resumen de verificacion
    if classes:
        print("  sample class:", json.dumps(classes[5], ensure_ascii=False)[:200])
    from collections import Counter
    print("  weapon types:", dict(Counter(w["weapon_type"] for w in weapons)))
    # Inspeccion cruda de Skills e Items para afinar el parser en el siguiente paso
    print("\n[Skills raw r2-4]")
    for r, vals in parse_two_col(wb["Skills"], 2)[:3]:
        print("  ", r, vals)
    print("[Items raw r1-4]")
    for r, vals in parse_two_col(wb["Items"], 1)[:4]:
        print("  ", r, vals)

if __name__ == "__main__":
    main()
