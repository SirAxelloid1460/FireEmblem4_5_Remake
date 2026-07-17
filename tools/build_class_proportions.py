#!/usr/bin/env python3
"""Proporcionalidad FE4 <-> FE5 por clase, a partir de los bloques versions de
data/general/classes.json.

Para cada clase con versiones FE4 y FE5:
  ratio_base[stat] = FE4.base[stat] / FE5.base[stat]   (si FE5>0)
  ratio_cap[stat]  = FE4.cap[stat]  / FE5.cap[stat]
  ratio_global     = sum(FE4.base) / sum(FE5.base)     (escala media de la clase)

Sirve para escalar stats de personajes entre la escala FE5 (baja) y la FE4/SAGA
(alta) en el modo combinado.  Salidas:
  data/general/class_proportions.json  (para el código)
  docs/FE4_FE5_Class_Proportions.xlsx  (referencia)
"""
import json, os

REPO = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
CLASSES = os.path.join(REPO, "data", "general", "classes.json")
OUT_JSON = os.path.join(REPO, "data", "general", "class_proportions.json")
OUT_XLSX = os.path.join(REPO, "docs", "FE4_FE5_Class_Proportions.xlsx")
S8 = ["HP", "STR", "MAG", "SKL", "SPD", "LCK", "DEF", "RES"]


def _ratio(a, b):
    return round(a / b, 3) if b else None


def compute():
    out = {}
    for c in json.load(open(CLASSES)):
        v = c.get("versions")
        if not v or "FE4" not in v or "FE5" not in v:
            continue
        b4 = v["FE4"].get("bases", {})
        b5 = v["FE5"].get("bases", {})
        c4 = v["FE4"].get("caps", {})
        c5 = v["FE5"].get("caps", {})
        if not (b4 and b5):
            continue
        rb = {s: _ratio(int(b4.get(s, 0)), int(b5.get(s, 0))) for s in S8}
        rc = {s: _ratio(int(c4.get(s, 0)), int(c5.get(s, 0))) for s in S8}
        sum4 = sum(int(b4.get(s, 0)) for s in S8)
        sum5 = sum(int(b5.get(s, 0)) for s in S8)
        out[c["id"]] = {
            "ratio_base": rb,
            "ratio_cap": rc,
            "ratio_global": _ratio(sum4, sum5),
            "sum_fe4": sum4,
            "sum_fe5": sum5,
        }
    return out


def write_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Proporciones"
    thin = Side(style="thin", color="C0C0C0"); box = Border(thin, thin, thin, thin)
    hdr = Font(bold=True, color="FFFFFF"); hf = PatternFill("solid", fgColor="374785")
    cen = Alignment(horizontal="center")
    cols = ["Clase"] + ["b." + s for s in S8] + ["Global"] + ["cap." + s for s in S8]
    for i, n in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=n); cell.font = hdr; cell.fill = hf
        cell.border = box; cell.alignment = cen
    r = 2
    for cid, d in sorted(data.items()):
        row = [cid] + [d["ratio_base"][s] for s in S8] + [d["ratio_global"]] \
            + [d["ratio_cap"][s] for s in S8]
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=val); cell.border = box; cell.alignment = cen
            if i == 1:
                cell.font = Font(bold=True)
        r += 1
    ws.column_dimensions["A"].width = 16
    for col in ws.iter_cols(min_col=2, max_col=len(cols)):
        ws.column_dimensions[col[0].column_letter].width = 7
    ws.freeze_panes = "B2"
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    data = compute()
    json.dump(data, open(OUT_JSON, "w"), indent=1, ensure_ascii=False)
    open(OUT_JSON, "a").write("\n")
    write_xlsx(data)
    print("classes with FE4/FE5 proportions:", len(data))
    print("wrote", OUT_JSON, "and", OUT_XLSX)
