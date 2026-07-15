#!/usr/bin/env python3
"""
asset_coverage.py

Genera una matriz de cobertura de assets de personajes y clases (map sprites +
animaciones de combate + retratos) para ir tachando a medida que se completan.

Salidas:
  - docs/ASSET_COVERAGE.xlsx   (Excel: hojas Resumen/Clases/Personajes/Problemas,
                                con colores, filtros y cabecera fija)  ← principal
  - docs/ASSET_COVERAGE.md     (Markdown, salvo --no-md)

Cruza:
  - data/general/classes.json  (clases: map_sprite_nid, combat_anim_nid, tier)
  - data/general/units.json    (personajes: nid, klass, gender)
  - assets/map_sprites/*-stand.png / *-move.png
  - assets/combat_anims/{NID}_{Variant}/*.png
  - assets/portraits/characters/{nid}.png

Uso:
  python tools/asset_coverage.py            # genera .xlsx (+ .md)
  python tools/asset_coverage.py --no-md    # solo .xlsx
  python tools/asset_coverage.py --no-xlsx  # solo .md
"""

import argparse
import json
import os
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MS_DIR = os.path.join(ROOT, "assets", "map_sprites")
CA_DIR = os.path.join(ROOT, "assets", "combat_anims")
PT_DIR = os.path.join(ROOT, "assets", "portraits", "characters")
OUT_MD = os.path.join(ROOT, "docs", "ASSET_COVERAGE.md")
OUT_XLSX = os.path.join(ROOT, "docs", "ASSET_COVERAGE.xlsx")

YES, NO, WARN, NA = "✅", "⬜", "⚠️", "—"


def load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def map_sprite_nids():
    stand, move = set(), set()
    if os.path.isdir(MS_DIR):
        for f in os.listdir(MS_DIR):
            if f.endswith("-stand.png"):
                stand.add(f[:-len("-stand.png")])
            elif f.endswith("-move.png"):
                move.add(f[:-len("-move.png")])
    return stand, move


def combat_anim_folders():
    out = {}
    if os.path.isdir(CA_DIR):
        for name in sorted(os.listdir(CA_DIR)):
            d = os.path.join(CA_DIR, name)
            if not os.path.isdir(d):
                continue
            weapons = set()
            for f in os.listdir(d):
                if f.endswith(".png") and f.startswith(name + "_"):
                    weapons.add(f[len(name) + 1:-4])
            out[name] = weapons
    return out


def portrait_nids():
    out = set()
    if os.path.isdir(PT_DIR):
        for f in os.listdir(PT_DIR):
            if f.endswith(".png"):
                out.add(f[:-4])
    return out


def mark(cond):
    return YES if cond else NO


def weapons_str(weapons):
    return ", ".join(sorted(weapons)) if weapons else "—"


# ══════════════════════════════════════════════════════════════════════════════
# Cálculo de datos
# ══════════════════════════════════════════════════════════════════════════════

def build_data():
    classes = load(os.path.join(ROOT, "data", "general", "classes.json"))
    units = load(os.path.join(ROOT, "data", "general", "units.json"))
    stand, move = map_sprite_nids()
    ca = combat_anim_folders()
    ca_names = set(ca.keys())
    portraits = portrait_nids()
    can_by_class = {str(c.get("id", "")): str(c.get("combat_anim_nid", "")) for c in classes}

    # Filas de clases.
    class_rows = []
    for c in sorted(classes, key=lambda x: (x.get("tier", 0), str(x.get("id", "")))):
        cid = str(c.get("id", ""))
        msn = str(c.get("map_sprite_nid", cid))
        can = str(c.get("combat_anim_nid", cid))
        gen = (can + "_Generic") in ca_names
        male = (can + "_Male") in ca_names
        fem = (can + "_Female") in ca_names
        # Generic y Male/Female son rutas EXCLUYENTES: si hay Generic no hacen
        # falta las de género (y viceversa). Lo no-necesario se marca "—".
        if gen:
            gcell, mcell, fcell = YES, NA, NA
            wsrc = ca.get(can + "_Generic", set())
        elif male or fem:
            gcell = NA
            mcell = YES if male else NO
            fcell = YES if fem else NO
            wsrc = ca.get(can + "_Male", set()) | ca.get(can + "_Female", set())
        else:
            gcell, mcell, fcell = NO, NA, NA   # sin cubrir → hace falta al menos Generic
            wsrc = set()
        class_rows.append([
            cid, c.get("tier", ""),
            mark(msn in stand), mark(msn in move),
            gcell, mcell, fcell,
            weapons_str(wsrc - {"Unarmed"}),
        ])

    # Filas de personajes.
    char_rows = []
    for u in sorted(units, key=lambda x: (str(x.get("klass", "")), str(x.get("nid", "")))):
        nid = str(u.get("nid", ""))
        klass = str(u.get("klass", ""))
        can = can_by_class.get(klass, "")
        prf_folder = next((f for f in sorted(ca_names) if f.endswith("_" + nid)), None)
        if prf_folder is None:
            anim_cell = NO
        elif can and prf_folder == can + "_" + nid:
            anim_cell = YES
        else:
            anim_cell = WARN
        char_rows.append([
            nid, klass, str(u.get("gender", "")),
            mark(nid in portraits), mark(nid in stand), anim_cell,
            weapons_str(ca.get(prf_folder, set()) - {"Unarmed"}) if prf_folder else "—",
        ])

    # Problemas (combat_anim_nid que rompe el resolver).
    problems = []
    for c in classes:
        can = str(c.get("combat_anim_nid", ""))
        if " " in can:
            problems.append(f"Clase '{c.get('id')}': combat_anim_nid = \"{can}\" tiene un "
                            "ESPACIO -> el resolver no lo encuentra. Quitar el espacio.")
    for u in units:
        nid = str(u.get("nid", ""))
        klass = str(u.get("klass", ""))
        can = can_by_class.get(klass, "")
        prf_folder = next((f for f in sorted(ca_names) if f.endswith("_" + nid)), None)
        if prf_folder and can and prf_folder != can + "_" + nid:
            problems.append(f"'{nid}' ({klass}): anim en {prf_folder}/ pero combat_anim_nid="
                            f"\"{can}\" -> el resolver busca {can}_{nid} (no existe). "
                            f"Renombrar la carpeta a {can}_{nid} o corregir combat_anim_nid.")

    summary = {
        "classes": len(classes),
        "classes_with_map": sum(1 for c in classes if str(c.get("map_sprite_nid", "")) in stand),
        "classes_with_generic_anim": sum(
            1 for c in classes
            if any((str(c.get("combat_anim_nid", "")) + s) in ca_names
                   for s in ("_Generic", "_Male", "_Female"))),
        "chars": len(units),
        "chars_with_portrait": sum(1 for u in units if str(u.get("nid", "")) in portraits),
        "anim_folders": len(ca),
    }
    return summary, class_rows, char_rows, problems


CLASS_HEADERS = ["Clase", "Tier", "Map stand", "Map move", "Anim Generic",
                 "Anim Male", "Anim Female", "Armas de la anim"]
CHAR_HEADERS = ["Personaje", "Clase", "Género", "Retrato", "Map propio",
                "Combat anim", "Armas de su anim"]
# Columnas (0-based) que llevan marca ✅/⬜/⚠️ (para colorear).
CLASS_MARK_COLS = [2, 3, 4, 5, 6]
CHAR_MARK_COLS = [3, 4, 5]


# ══════════════════════════════════════════════════════════════════════════════
# Escritura Markdown
# ══════════════════════════════════════════════════════════════════════════════

def write_md(summary, class_rows, char_rows, problems):
    L = []
    L.append("# Cobertura de assets — personajes y clases\n")
    L.append("> Generado por `tools/asset_coverage.py`.\n")
    L.append(f"> Leyenda: {YES} = tenemos · {NO} = falta · {WARN} = existe pero no "
             "coincide con `combat_anim_nid` (el juego no lo encuentra) · — = N/A\n")
    L.append("\n## Resumen\n")
    L.append(f"- Clases: **{summary['classes']}** · con map sprite: "
             f"**{summary['classes_with_map']}** · con combat anim (Generic o género): "
             f"**{summary['classes_with_generic_anim']}**")
    L.append(f"- Personajes: **{summary['chars']}** · con retrato: "
             f"**{summary['chars_with_portrait']}**")
    L.append(f"- Carpetas de combat anim: **{summary['anim_folders']}**\n")

    L.append("\n## Clases\n")
    L.append("| " + " | ".join(CLASS_HEADERS) + " |")
    L.append("|" + "|".join(["---"] * len(CLASS_HEADERS)) + "|")
    for r in class_rows:
        L.append("| " + " | ".join(str(x) for x in r) + " |")

    L.append("\n## Personajes\n")
    L.append("| " + " | ".join(CHAR_HEADERS) + " |")
    L.append("|" + "|".join(["---"] * len(CHAR_HEADERS)) + "|")
    for r in char_rows:
        L.append("| " + " | ".join(str(x) for x in r) + " |")

    if problems:
        L.append("\n## ⚠️ Problemas detectados (el juego no encuentra estos anims)\n")
        for p in problems:
            L.append("- " + p)

    os.makedirs(os.path.dirname(OUT_MD), exist_ok=True)
    with open(OUT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print("Escrito", os.path.relpath(OUT_MD, ROOT))


# ══════════════════════════════════════════════════════════════════════════════
# Escritura Excel
# ══════════════════════════════════════════════════════════════════════════════

def write_xlsx(summary, class_rows, char_rows, problems):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("AVISO: openpyxl no instalado — omito el .xlsx "
              "(instala con: py -m pip install openpyxl)")
        return

    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    fills = {YES: PatternFill("solid", fgColor="C6EFCE"),
             NO: PatternFill("solid", fgColor="F2F2F2"),
             WARN: PatternFill("solid", fgColor="FFEB9C"),
             NA: PatternFill("solid", fgColor="FFFFFF")}   # neutro (no aplica)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def add_table(title, headers, rows, mark_cols, widths):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for j, _ in enumerate(headers, start=1):
            c = ws.cell(row=1, column=j)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border
        for row in rows:
            ws.append(list(row))
        # Estilo de celdas de datos.
        for i in range(2, len(rows) + 2):
            for j in range(1, len(headers) + 1):
                c = ws.cell(row=i, column=j)
                c.border = border
                if (j - 1) in mark_cols:
                    c.alignment = center
                    c.fill = fills.get(str(c.value), fills[NO])
        # Anchos, freeze de cabecera, autofiltro.
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), len(rows) + 1)
        return ws

    # Hoja Resumen.
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Cobertura de assets — resumen"
    ws["A1"].font = Font(bold=True, size=14)
    data = [
        ("Clases", summary["classes"]),
        ("  con map sprite", summary["classes_with_map"]),
        ("  con combat anim (Generic o género)", summary["classes_with_generic_anim"]),
        ("Personajes", summary["chars"]),
        ("  con retrato", summary["chars_with_portrait"]),
        ("Carpetas de combat anim", summary["anim_folders"]),
        ("Problemas detectados", len(problems)),
    ]
    for i, (k, v) in enumerate(data, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.cell(row=len(data) + 4, column=1,
            value=f"Leyenda: {YES}=tenemos  {NO}=falta  {WARN}=existe pero no resuelve  —=N/A")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12

    add_table("Clases", CLASS_HEADERS, class_rows, CLASS_MARK_COLS,
              [18, 6, 11, 11, 13, 12, 13, 34])
    add_table("Personajes", CHAR_HEADERS, char_rows, CHAR_MARK_COLS,
              [16, 16, 9, 10, 12, 14, 30])

    wsp = wb.create_sheet("Problemas")
    wsp.append(["#", "Problema"])
    for j in (1, 2):
        wsp.cell(row=1, column=j).fill = header_fill
        wsp.cell(row=1, column=j).font = header_font
    for i, p in enumerate(problems, start=1):
        wsp.cell(row=i + 1, column=1, value=i)
        wsp.cell(row=i + 1, column=2, value=p)
    wsp.column_dimensions["A"].width = 5
    wsp.column_dimensions["B"].width = 120
    wsp.freeze_panes = "A2"

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    print("Escrito", os.path.relpath(OUT_XLSX, ROOT))


def main():
    ap = argparse.ArgumentParser(description="Matriz de cobertura de assets (Excel + Markdown).")
    ap.add_argument("--no-md", action="store_true", help="No generar el .md.")
    ap.add_argument("--no-xlsx", action="store_true", help="No generar el .xlsx.")
    args = ap.parse_args()

    summary, class_rows, char_rows, problems = build_data()
    if not args.no_xlsx:
        write_xlsx(summary, class_rows, char_rows, problems)
    if not args.no_md:
        write_md(summary, class_rows, char_rows, problems)
    print(f"  clases={summary['classes']} (map={summary['classes_with_map']}, "
          f"anim_gen={summary['classes_with_generic_anim']}) "
          f"personajes={summary['chars']} (retrato={summary['chars_with_portrait']}) "
          f"problemas={len(problems)}")


if __name__ == "__main__":
    main()
